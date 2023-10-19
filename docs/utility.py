import os
import pandas as pd
from openpyxl import *




def set_str(value):
    if value == None or str(value) == str(""):
        v = ""
    else:
        v = str(value)
    return v

def set_int(value):
    if value != value or str(value) == str(""):
        v = ""
    else:
        v = int(value)
    return v

def set_float(value):
    if value == None or str(value) == str(""):
            v = ""
    else:
        v = float(value)
    return v


def get_excel_sheets(xls_file):
    sheets = []
    if not os.path.isfile(xls_file):
        return sheets
    
    sheets = list(pd.read_excel(xls_file, sheet_name=None, engine='openpyxl'))
    return sheets

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    # base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    base_path = os.getcwd()
    return os.path.join(base_path, relative_path)

def check_tips(tips_str):
    '''
    将;分割的字符串转换为列表
    '''
    if type(tips_str) != type(""):
        return [""]
    return tips_str.split(";")

# pandas to_csv float格式任意精度，无需工程格式化
def export_formatted(df, csv_path, cols=None):

    # By default, format all columns in df
    print(df)
    if cols==None:
        cols = df.columns
    print(cols)
    # Change columns to strings with 0's stripped as desired
    for c in cols:
        if type(c) is not str:        
            df[c] = df[c].map('{:,.15f}'.format).str.rstrip('0')

    # export
    df.to_csv(csv_path)

def get_excel_sheets(xls_file):
    sheets = []
    if not os.path.isfile(xls_file):
        return sheets
    
    sheets = list(pd.read_excel(xls_file, sheet_name=None, engine='openpyxl'))
    return sheets

def check_tips(tips_str):
    '''
    将;分割的字符串转换为列表
    '''
    if type(tips_str) != type(""):
        return [""]
    return tips_str.split(";")

def get_tips(xls_file_path):
    '''
    获取当前数据表对应标签列
    '''
    if "tags@" not in get_excel_sheets(xls_file_path):
        raise Exception(xls_file_path + '不含有叫做tags@的sheet！！！')
    tips_df = pd.read_excel(xls_file_path,sheet_name="tags@", engine='openpyxl', keep_default_na=False)
    tips_dict = tips_df.to_dict()
    return list(tips_dict["列标签"].values())

class excel():
    def __init__(self,file):
        self.file = file
        self.wb = load_workbook(self.file)
        sheets = self.wb.get_sheet_names()
        self.sheet = sheets[0]
        self.ws = self.wb[self.sheet]
        
    #获取表格的总行数和总列数
    def getRowsClosNum(self):
        rows = self.ws.max_row
        columns = self.ws.max_column
        return rows,columns
    
    #获取某个单元格的值
    def getCellValue(self,row,column):       
        cellvalue = self.ws.cell(row=row,column=column).value
        return cellvalue
    
    #获取某列的全部值
    def getColValues(self,column):
        rows = self.ws.max_row
        columndata=[]
        for i in range(1,rows+1):
            cellvalue = self.ws.cell(row=i,column=column).value
            columndata.append(cellvalue)
        return columndata
    
    
    #获取某行全部值
    def getRowValues(self,row):
        columns = self.ws.max_column
        rowdata=[]
        for i in range(1,columns+1):
            cellvalue = self.ws.cell(row=row,column=i).value
            rowdata.append(cellvalue)
        return rowdata
    
    #设置某个单元格的值
    def setCellValue(self,row,colunm,cellvalue):
        try:
            self.ws.cell(row=row,column=colunm).value=cellvalue
            self.wb.save(self.file)
        except:
            self.ws.cell(row=row,column=colunm).value="writefail"
            self.wb.save(self.file)